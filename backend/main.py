from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
import sqlite3
import threading
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

DB_PATH = Path(__file__).parent / "docfind.db"
SUPPORTED_EXTENSIONS = {".pdf", ".xls", ".xlsx", ".txt", ".csv", ".log"}
DEFAULT_EXCLUDES = ["~$*", "*.tmp", "*.bak", ".*", "__pycache__"]


class SearchRequest(BaseModel):
    query: str
    sort: str = "relevance"
    view: str = "list"
    limit: int = 100
    offset: int = 0
    file_types: list[str] = Field(default_factory=list)
    folders: list[str] = Field(default_factory=list)
    tags: list[str] = Field(default_factory=list)
    modified_from: str | None = None
    modified_to: str | None = None
    size_min_kb: int | None = None
    size_max_kb: int | None = None
    sheet: str | None = None
    column: str | None = None


class IndexRequest(BaseModel):
    folders: list[str]
    recursive: bool = True


class TagRequest(BaseModel):
    tag: str


class NoteRequest(BaseModel):
    note: str


class SaveSearchRequest(BaseModel):
    name: str
    query: str
    filters: dict[str, Any] = Field(default_factory=dict)


@dataclass
class ExtractedChunk:
    ref: str
    content: str


app = FastAPI(title="DocFind")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

INDEX_STATUS = {
    "running": False,
    "started_at": None,
    "finished_at": None,
    "processed": 0,
    "total": 0,
    "last_error": None,
}


# ------------------------------
# DB utilities
# ------------------------------
def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def setup_db() -> None:
    conn = get_conn()
    cur = conn.cursor()
    cur.executescript(
        """
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY,
            path TEXT UNIQUE,
            name TEXT,
            extension TEXT,
            size INTEGER,
            created_at TEXT,
            modified_at TEXT,
            author TEXT,
            page_count INTEGER,
            sheet_count INTEGER,
            password_protected INTEGER DEFAULT 0,
            index_status TEXT DEFAULT 'indexed',
            note TEXT DEFAULT '',
            favorite INTEGER DEFAULT 0,
            indexed_at TEXT,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS file_chunks (
            id INTEGER PRIMARY KEY,
            file_id INTEGER NOT NULL,
            ref TEXT NOT NULL,
            content TEXT NOT NULL,
            FOREIGN KEY(file_id) REFERENCES files(id) ON DELETE CASCADE
        );

        CREATE VIRTUAL TABLE IF NOT EXISTS file_chunks_fts USING fts5(
            content,
            file_id UNINDEXED,
            ref UNINDEXED,
            tokenize='porter unicode61'
        );

        CREATE TABLE IF NOT EXISTS tags (
            id INTEGER PRIMARY KEY,
            file_id INTEGER NOT NULL,
            tag TEXT NOT NULL,
            UNIQUE(file_id, tag),
            FOREIGN KEY(file_id) REFERENCES files(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS folders (
            id INTEGER PRIMARY KEY,
            path TEXT UNIQUE,
            enabled INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS saved_searches (
            id INTEGER PRIMARY KEY,
            name TEXT,
            query TEXT,
            filters_json TEXT,
            created_at TEXT
        );

        CREATE TABLE IF NOT EXISTS search_history (
            id INTEGER PRIMARY KEY,
            query TEXT,
            created_at TEXT
        );

        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );
        """
    )
    conn.commit()
    conn.close()


def now_iso() -> str:
    return datetime.utcnow().isoformat()


# ------------------------------
# Parsers
# ------------------------------
def extract_text(path: Path) -> tuple[list[ExtractedChunk], dict[str, Any]]:
    ext = path.suffix.lower()
    if ext in {".txt", ".csv", ".log"}:
        return extract_plain_text(path), {}
    if ext == ".pdf":
        return extract_pdf(path)
    if ext in {".xls", ".xlsx"}:
        return extract_excel(path)
    return [], {}


def extract_plain_text(path: Path) -> list[ExtractedChunk]:
    try:
        text = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        text = path.read_text(encoding="latin-1", errors="ignore")

    chunks: list[ExtractedChunk] = []
    for idx, line in enumerate(text.splitlines()):
        if line.strip():
            chunks.append(ExtractedChunk(ref=f"line:{idx + 1}", content=line.strip()))
    if not chunks and text.strip():
        chunks.append(ExtractedChunk(ref="line:1", content=text.strip()))
    return chunks


def extract_pdf(path: Path) -> tuple[list[ExtractedChunk], dict[str, Any]]:
    metadata: dict[str, Any] = {}
    try:
        from pypdf import PdfReader
    except Exception:
        # Graceful fallback when optional dependency is missing.
        return [ExtractedChunk(ref="page:1", content="[Unable to parse PDF: install pypdf]")], {"parse_error": "pypdf missing"}

    reader = PdfReader(str(path))
    if reader.is_encrypted:
        return [], {"password_protected": 1}

    chunks: list[ExtractedChunk] = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if not text.strip():
            # OCR fallback placeholder
            text = "[Scanned page detected: install OCR dependencies for text extraction]"
        chunks.append(ExtractedChunk(ref=f"page:{i + 1}", content=text.strip()))

    info = reader.metadata or {}
    metadata["author"] = str(info.get("/Author", "")) if info else ""
    metadata["page_count"] = len(reader.pages)
    return chunks, metadata


def extract_excel(path: Path) -> tuple[list[ExtractedChunk], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return [ExtractedChunk(ref="sheet:Unknown,row:0", content="[Unable to parse Excel: install openpyxl]")], {"parse_error": "openpyxl missing"}

    wb_values = load_workbook(path, data_only=True, read_only=True)
    chunks: list[ExtractedChunk] = []
    for sheet in wb_values.sheetnames:
        ws = wb_values[sheet]
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = ["" if v is None else str(v) for v in row]
            if row_idx == 1:
                headers = values
            joined = " | ".join(values).strip(" |")
            if joined:
                ref = f"sheet:{sheet},row:{row_idx}"
                chunks.append(ExtractedChunk(ref=ref, content=joined))
                # enrich searchable content with column header context
                if headers and row_idx > 1:
                    enriched = " ".join(f"{headers[i]}:{values[i]}" for i in range(min(len(headers), len(values))) if headers[i])
                    if enriched:
                        chunks.append(ExtractedChunk(ref=f"sheet:{sheet},row:{row_idx},headers", content=enriched))

    return chunks, {"sheet_count": len(wb_values.sheetnames)}


def should_exclude(path: Path, patterns: list[str]) -> bool:
    from fnmatch import fnmatch

    parts = list(path.parts)
    return any(fnmatch(path.name, p) for p in patterns) or any(any(fnmatch(part, p) for p in patterns) for part in parts)


# ------------------------------
# Indexing
# ------------------------------
def scan_files(folders: list[str], recursive: bool = True) -> list[Path]:
    files: list[Path] = []
    for folder in folders:
        base = Path(folder).expanduser()
        if not base.exists():
            continue
        it = base.rglob("*") if recursive else base.glob("*")
        for p in it:
            if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS:
                files.append(p)
    return files


def clear_file_index(cur: sqlite3.Cursor, file_id: int) -> None:
    cur.execute("DELETE FROM file_chunks WHERE file_id=?", (file_id,))
    cur.execute("DELETE FROM file_chunks_fts WHERE file_id=?", (file_id,))


def index_file(cur: sqlite3.Cursor, path: Path) -> None:
    st = path.stat()
    created = datetime.utcfromtimestamp(st.st_ctime).isoformat()
    modified = datetime.utcfromtimestamp(st.st_mtime).isoformat()

    chunks, metadata = extract_text(path)

    cur.execute(
        """
        INSERT INTO files(path, name, extension, size, created_at, modified_at, author, page_count, sheet_count,
                          password_protected, indexed_at, updated_at, index_status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(path) DO UPDATE SET
            name=excluded.name,
            extension=excluded.extension,
            size=excluded.size,
            created_at=excluded.created_at,
            modified_at=excluded.modified_at,
            author=excluded.author,
            page_count=excluded.page_count,
            sheet_count=excluded.sheet_count,
            password_protected=excluded.password_protected,
            indexed_at=excluded.indexed_at,
            updated_at=excluded.updated_at,
            index_status=excluded.index_status
        """,
        (
            str(path),
            path.name,
            path.suffix.lower(),
            st.st_size,
            created,
            modified,
            metadata.get("author", ""),
            metadata.get("page_count"),
            metadata.get("sheet_count"),
            int(bool(metadata.get("password_protected", 0))),
            now_iso(),
            now_iso(),
            "needs_password" if metadata.get("password_protected") else "indexed",
        ),
    )

    file_id = cur.execute("SELECT id FROM files WHERE path=?", (str(path),)).fetchone()[0]
    clear_file_index(cur, file_id)

    if metadata.get("password_protected"):
        return

    for ch in chunks:
        cur.execute("INSERT INTO file_chunks(file_id, ref, content) VALUES (?, ?, ?)", (file_id, ch.ref, ch.content))
        cur.execute("INSERT INTO file_chunks_fts(content, file_id, ref) VALUES (?, ?, ?)", (ch.content, file_id, ch.ref))


def run_full_index(folders: list[str], recursive: bool = True) -> None:
    INDEX_STATUS.update(
        {
            "running": True,
            "started_at": now_iso(),
            "finished_at": None,
            "processed": 0,
            "last_error": None,
        }
    )
    conn = get_conn()
    cur = conn.cursor()
    try:
        files = scan_files(folders, recursive=recursive)
        INDEX_STATUS["total"] = len(files)
        cur.executemany("INSERT OR IGNORE INTO folders(path, enabled) VALUES (?,1)", [(f,) for f in folders])

        for p in files:
            try:
                index_file(cur, p)
                INDEX_STATUS["processed"] += 1
            except Exception as e:
                INDEX_STATUS["last_error"] = str(e)
        conn.commit()

        known_paths = {str(p) for p in files}
        rows = cur.execute("SELECT id, path FROM files").fetchall()
        for row in rows:
            if row["path"] not in known_paths:
                cur.execute("DELETE FROM files WHERE id=?", (row["id"],))
        conn.commit()
    finally:
        conn.close()
        INDEX_STATUS["running"] = False
        INDEX_STATUS["finished_at"] = now_iso()


# ------------------------------
# Search
# ------------------------------
def normalize_query(query: str) -> str:
    q = query.strip()
    if not q:
        return q

    # Proximity syntax: budget NEAR/5 2024
    q = re.sub(r"\bnear\s*/\s*(\d+)\b", r"NEAR/\1", q, flags=re.IGNORECASE)
    # Boolean uppercase for readability; SQLite accepts uppercase forms in MATCH grammar.
    q = re.sub(r"\band\b", "AND", q, flags=re.IGNORECASE)
    q = re.sub(r"\bor\b", "OR", q, flags=re.IGNORECASE)
    q = re.sub(r"\bnot\b", "NOT", q, flags=re.IGNORECASE)
    # wildcard already supported via * ; map ? to * for best-effort
    q = q.replace("?", "*")
    return q


def apply_filters(base_sql: str, request: SearchRequest) -> tuple[str, list[Any]]:
    where = []
    params: list[Any] = []
    if request.file_types:
        where.append(f"f.extension IN ({','.join('?' for _ in request.file_types)})")
        params.extend([ft.lower() if ft.startswith(".") else f".{ft.lower()}" for ft in request.file_types])
    if request.folders:
        folder_where = []
        for folder in request.folders:
            folder_where.append("f.path LIKE ?")
            params.append(f"{folder}%")
        where.append("(" + " OR ".join(folder_where) + ")")
    if request.tags:
        where.append("EXISTS (SELECT 1 FROM tags t WHERE t.file_id=f.id AND t.tag IN ({}) )".format(",".join("?" for _ in request.tags)))
        params.extend(request.tags)
    if request.modified_from:
        where.append("f.modified_at >= ?")
        params.append(request.modified_from)
    if request.modified_to:
        where.append("f.modified_at <= ?")
        params.append(request.modified_to)
    if request.size_min_kb is not None:
        where.append("f.size >= ?")
        params.append(request.size_min_kb * 1024)
    if request.size_max_kb is not None:
        where.append("f.size <= ?")
        params.append(request.size_max_kb * 1024)
    if request.sheet:
        where.append("fc.ref LIKE ?")
        params.append(f"sheet:{request.sheet}%")
    if request.column:
        where.append("fc.content LIKE ?")
        params.append(f"%{request.column}:%")

    if where:
        return base_sql + " WHERE " + " AND ".join(where), params
    return base_sql, params


def build_snippet(text: str, terms: list[str], radius: int = 80) -> str:
    lower = text.lower()
    pos = min([lower.find(t.lower()) for t in terms if t and lower.find(t.lower()) != -1] + [0])
    start = max(0, pos - radius)
    end = min(len(text), pos + radius)
    snippet = text[start:end].replace("\n", " ")
    for t in sorted(set(terms), key=len, reverse=True):
        if t:
            snippet = re.sub(re.escape(t), lambda m: f"<mark>{m.group(0)}</mark>", snippet, flags=re.IGNORECASE)
    return snippet


def score_result(row: sqlite3.Row, query_terms: list[str]) -> float:
    content = row["content"].lower()
    score = 0.0
    for term in query_terms:
        if not term:
            continue
        count = content.count(term.lower())
        score += count * 2
        if row["name"].lower().startswith(term.lower()):
            score += 1.5
    if row["extension"] == ".pdf":
        score += 0.1
    return round(score, 3)


@app.get("/api/search_history")
def search_history() -> dict[str, Any]:
    conn = get_conn()
    rows = conn.execute("SELECT query, created_at FROM search_history ORDER BY id DESC LIMIT 20").fetchall()
    conn.close()
    return {"history": [dict(r) for r in rows]}


@app.delete("/api/search_history")
def clear_search_history() -> dict[str, str]:
    conn = get_conn()
    conn.execute("DELETE FROM search_history")
    conn.commit()
    conn.close()
    return {"status": "cleared"}


@app.post("/api/search")
def search(request: SearchRequest) -> dict[str, Any]:
    conn = get_conn()
    q = normalize_query(request.query)
    if q:
        conn.execute("INSERT INTO search_history(query, created_at) VALUES (?, ?)", (request.query, now_iso()))

    base_sql = (
        """
        SELECT f.id AS file_id, f.name, f.path, f.extension, f.size, f.modified_at, f.author,
               fc.ref, fc.content
        FROM file_chunks_fts fts
        JOIN files f ON f.id = fts.file_id
        JOIN file_chunks fc ON fc.file_id=f.id AND fc.ref=fts.ref AND fc.content=fts.content
        """
    )

    sql, params = apply_filters(base_sql, request)
    if q:
        sql += (" AND " if " WHERE " in sql else " WHERE ") + "fts.content MATCH ?"
        params.append(q)

    rows = conn.execute(sql, params).fetchall()

    # Fuzzy fallback for typo-like queries
    if not rows and q and " " not in q and '"' not in q and "*" not in q:
        vocab = conn.execute("SELECT content FROM file_chunks LIMIT 10000").fetchall()
        corpus = set()
        for r in vocab:
            corpus.update(re.findall(r"[A-Za-z0-9_]{3,}", r["content"].lower()))
        import difflib

        alts = difflib.get_close_matches(q.lower(), list(corpus), n=3, cutoff=0.8)
        if alts:
            alt_q = " OR ".join(alts)
            rows = conn.execute(
                base_sql + (" WHERE " + "fts.content MATCH ?"),
                [alt_q],
            ).fetchall()

    grouped: dict[int, dict[str, Any]] = {}
    raw_terms = [t.strip('"') for t in re.split(r"\s+|AND|OR|NOT", request.query) if t and t.upper() not in {"AND", "OR", "NOT"}]

    for row in rows:
        item = grouped.setdefault(
            row["file_id"],
            {
                "file_id": row["file_id"],
                "name": row["name"],
                "path": row["path"],
                "extension": row["extension"],
                "size": row["size"],
                "modified_at": row["modified_at"],
                "author": row["author"],
                "matches": [],
                "match_count": 0,
                "score": 0.0,
            },
        )
        snip = build_snippet(row["content"], raw_terms)
        item["matches"].append({"ref": row["ref"], "snippet": snip, "content": row["content"]})
        item["match_count"] += 1
        item["score"] += score_result(row, raw_terms)

    results = list(grouped.values())

    if request.sort == "name_asc":
        results.sort(key=lambda r: r["name"].lower())
    elif request.sort == "date_modified":
        results.sort(key=lambda r: r["modified_at"], reverse=True)
    elif request.sort == "file_size":
        results.sort(key=lambda r: r["size"], reverse=True)
    else:
        results.sort(key=lambda r: r["score"], reverse=True)

    facets = compute_facets(conn, results)
    total = len(results)
    paged = results[request.offset : request.offset + request.limit]

    conn.commit()
    conn.close()
    return {"total": total, "results": paged, "facets": facets}


def compute_facets(conn: sqlite3.Connection, results: list[dict[str, Any]]) -> dict[str, Any]:
    if not results:
        return {"types": [], "folders": [], "tags": []}

    file_ids = [r["file_id"] for r in results]
    placeholders = ",".join("?" for _ in file_ids)
    type_rows = conn.execute(
        f"SELECT extension, COUNT(*) as c FROM files WHERE id IN ({placeholders}) GROUP BY extension", file_ids
    ).fetchall()
    folder_counter = Counter(str(Path(r["path"]).parent) for r in results)
    tag_rows = conn.execute(
        f"SELECT tag, COUNT(*) c FROM tags WHERE file_id IN ({placeholders}) GROUP BY tag", file_ids
    ).fetchall()

    return {
        "types": [dict(r) for r in type_rows],
        "folders": [{"folder": k, "count": v} for k, v in folder_counter.most_common(20)],
        "tags": [dict(r) for r in tag_rows],
    }


@app.post("/api/index")
def trigger_index(request: IndexRequest) -> dict[str, Any]:
    if INDEX_STATUS["running"]:
        raise HTTPException(status_code=409, detail="Index already running")

    thread = threading.Thread(target=run_full_index, args=(request.folders, request.recursive), daemon=True)
    thread.start()
    return {"status": "started"}


@app.get("/api/index/status")
def index_status() -> dict[str, Any]:
    conn = get_conn()
    total = conn.execute("SELECT COUNT(*) c FROM files").fetchone()["c"]
    index_size = DB_PATH.stat().st_size if DB_PATH.exists() else 0
    last_index = conn.execute("SELECT MAX(indexed_at) AS t FROM files").fetchone()["t"]
    conn.close()
    return {
        **INDEX_STATUS,
        "total_files_indexed": total,
        "index_size_bytes": index_size,
        "last_indexed_time": last_index,
    }


@app.post("/api/index/clear")
def clear_index() -> dict[str, str]:
    conn = get_conn()
    conn.executescript(
        """
        DELETE FROM file_chunks;
        DELETE FROM file_chunks_fts;
        DELETE FROM files;
        DELETE FROM tags;
        """
    )
    conn.commit()
    conn.close()
    return {"status": "cleared"}


@app.post("/api/files/{file_id}/tag")
def add_tag(file_id: int, request: TagRequest) -> dict[str, str]:
    conn = get_conn()
    conn.execute("INSERT OR IGNORE INTO tags(file_id, tag) VALUES (?, ?)", (file_id, request.tag.strip()))
    conn.commit()
    conn.close()
    return {"status": "ok"}


@app.post("/api/files/{file_id}/note")
def save_note(file_id: int, request: NoteRequest) -> dict[str, str]:
    conn = get_conn()
    conn.execute("UPDATE files SET note=? WHERE id=?", (request.note, file_id))
    conn.commit()
    conn.close()
    return {"status": "ok"}


@app.post("/api/saved_searches")
def save_search(req: SaveSearchRequest) -> dict[str, str]:
    conn = get_conn()
    conn.execute(
        "INSERT INTO saved_searches(name, query, filters_json, created_at) VALUES (?, ?, ?, ?)",
        (req.name, req.query, json.dumps(req.filters), now_iso()),
    )
    conn.commit()
    conn.close()
    return {"status": "saved"}


@app.get("/api/saved_searches")
def list_saved_searches() -> dict[str, Any]:
    conn = get_conn()
    rows = conn.execute("SELECT id, name, query, filters_json, created_at FROM saved_searches ORDER BY id DESC").fetchall()
    conn.close()
    return {"items": [dict(r) for r in rows]}


@app.get("/api/export")
def export_results_csv(q: str = "") -> FileResponse:
    req = SearchRequest(query=q)
    data = search(req)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["file_name", "path", "match_count", "matched_snippet", "date_modified"])
    writer.writeheader()
    for row in data["results"]:
        snippet = row["matches"][0]["snippet"] if row["matches"] else ""
        writer.writerow(
            {
                "file_name": row["name"],
                "path": row["path"],
                "match_count": row["match_count"],
                "matched_snippet": re.sub("<[^>]+>", "", snippet),
                "date_modified": row["modified_at"],
            }
        )

    temp_path = Path(__file__).parent / "export_results.csv"
    temp_path.write_text(output.getvalue(), encoding="utf-8")
    return FileResponse(str(temp_path), media_type="text/csv", filename="docfind_results.csv")


@app.get("/api/analytics")
def analytics() -> dict[str, Any]:
    conn = get_conn()
    by_type = conn.execute("SELECT extension, COUNT(*) c FROM files GROUP BY extension").fetchall()
    top_terms = conn.execute(
        "SELECT query, COUNT(*) c FROM search_history GROUP BY query ORDER BY c DESC LIMIT 10"
    ).fetchall()
    recent_activity = conn.execute(
        "SELECT query, created_at FROM search_history ORDER BY id DESC LIMIT 25"
    ).fetchall()
    conn.close()
    return {
        "files_by_type": [dict(r) for r in by_type],
        "top_terms": [dict(r) for r in top_terms],
        "recent_activity": [dict(r) for r in recent_activity],
        "index_storage_bytes": DB_PATH.stat().st_size if DB_PATH.exists() else 0,
    }


@app.get("/api/open_file")
def open_file(path: str) -> dict[str, Any]:
    # Best effort helper for desktop runtime (Electron shell.openPath equivalent).
    return {"path": path, "message": "Open externally handled by shell integration in desktop wrapper."}


@app.get("/", response_class=HTMLResponse)
def root() -> str:
    html = (Path(__file__).parent.parent / "frontend" / "index.html").read_text(encoding="utf-8")
    return html


app.mount("/static", StaticFiles(directory=Path(__file__).parent.parent / "frontend"), name="static")


setup_db()

if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
